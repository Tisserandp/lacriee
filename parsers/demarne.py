"""
Parser pour les fichiers Excel Demarne.

Structure du document:
- Excel avec cellules fusionnées pour Calibre et Origine
- Catégories très détaillées combinant espèce + origine + type production
- Variantes contenant découpes, états et préparations

Attributs extraits:
- Categorie: Catégorie détaillée du produit
- Variante: Variante/préparation du produit
- Methode_Peche: LIGNE, CHALUT, CASIER, etc. (extrait de Variante/Label)
- Label: Certifications (MSC, BIO, ASC, Label Rouge, IGP, AOP)
- Calibre: Tailles variées (plages, N°, descriptifs)
- Origine: Pays, zones FAO, régions

Harmonisation:
- Ce parseur supporte l'harmonisation via services/harmonize.py
- Utiliser parse(file_bytes, harmonize=True) pour obtenir des attributs normalisés
- Voir docs/harmonisation_attributs.md pour les règles de normalisation

Spécificités Demarne:
- La catégorie contient souvent espèce + type_production + origine
  Ex: "SAUMON SUPÉRIEUR NORVÈGE" → categorie=SAUMON, qualite=SUP, origine=NORVÈGE
- Les huîtres par marque sont normalisées vers HUITRES
- Les labels MSC, BIO, ASC, Label Rouge, IGP, AOP sont extraits
- Les trims saumon (Trim B, D, E) sont extraits vers le champ trim
"""
import re
import io
import logging
from datetime import datetime, date
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Import conditionnel pour éviter les erreurs si harmonize.py n'existe pas encore
try:
    from services.harmonize import harmonize_products
    HARMONIZE_AVAILABLE = True
except ImportError:
    HARMONIZE_AVAILABLE = False

# Import de sanitize_for_json depuis parsers.utils
try:
    from parsers.utils import sanitize_for_json
except ImportError:
    def sanitize_for_json(df: pd.DataFrame) -> list[dict]:
        """Fallback si parsers.utils n'existe pas."""
        df = df.replace([float("inf"), float("-inf"), float('inf'), -float('inf')], None)
        df = df.where(pd.notnull(df), None)
        clean_data = []
        for _, row in df.iterrows():
            clean_row = {}
            for col, val in row.items():
                if isinstance(val, float) and (pd.isna(val) or val in [float("inf"), float("-inf")]):
                    clean_row[col] = None
                elif isinstance(val, str) and val.strip() == "":
                    clean_row[col] = None
                else:
                    clean_row[col] = val
            clean_data.append(clean_row)
        return clean_data

# Import optionnel de BigQuery pour la lookup table
try:
    from services.bigquery import get_bigquery_client
    BIGQUERY_AVAILABLE = True
except ImportError:
    BIGQUERY_AVAILABLE = False


def get_bigquery_client_internal():
    """Wrapper pour obtenir le client BigQuery."""
    if BIGQUERY_AVAILABLE:
        return get_bigquery_client()
    else:
        raise ImportError(
            "Le module services.bigquery n'est pas disponible. "
            "Vérifiez que services/bigquery.py existe."
        )


def extract_date_from_excel_header(file_input, date_fallback: Optional[str] = None) -> Optional[str]:
    """
    Extrait la date depuis les métadonnées Excel (header).
    Si aucune date n'est trouvée, utilise date_fallback si fourni.
    Retourne None seulement si aucune date n'est trouvée ET aucun fallback fourni.
    """
    # Gérer l'entrée bytes
    if isinstance(file_input, bytes):
        wb_input = io.BytesIO(file_input)
    else:
        wb_input = file_input

    try:
        wb = load_workbook(wb_input, data_only=True)
        sheet = wb.active
        header = sheet.oddHeader.center.text if sheet.oddHeader and sheet.oddHeader.center else ""
        match = re.search(r"\b(\d{2})/(\d{2})/(\d{4})\b", header)
        if match:
            day, month, year = match.groups()
            return date(int(year), int(month), int(day)).isoformat()
    except Exception as e:
        logger.debug(f"Erreur lors de la lecture du header Excel: {e}")

    # Fallback: utiliser la date fournie par l'utilisateur
    if date_fallback:
        # Valider le format de la date fallback (YYYY-MM-DD)
        try:
            parsed_date = datetime.strptime(date_fallback, "%Y-%m-%d").date()
            return parsed_date.isoformat()
        except ValueError:
            # Essayer aussi le format DD/MM/YYYY
            try:
                parsed_date = datetime.strptime(date_fallback, "%d/%m/%Y").date()
                return parsed_date.isoformat()
            except ValueError:
                raise ValueError(f"Format de date invalide pour le fallback: '{date_fallback}'. Utilisez YYYY-MM-DD ou DD/MM/YYYY")

    return None


def split_fr_en(text: str):
    """
    Sépare un texte français/anglais séparé par "/"
    Ex: "SAUMON SUPÉRIEUR / SUPERIOR SALMON" -> ("SAUMON SUPÉRIEUR", "SUPERIOR SALMON")
    Ex: "Dorade Royale/ Gilthead bream" -> ("Dorade Royale", "Gilthead bream")
    Ex: "Dorade Royale / Gilthead bream" -> ("Dorade Royale", "Gilthead bream")
    Gère aussi les cas où le "/" est dans un mot (ex: "Trim B/D" ne sera pas split)
    """
    if not text or not isinstance(text, str):
        return text, None

    text = text.strip()

    # Chercher un "/" avec au moins un espace d'un côté, ou "/" suivi d'une majuscule
    # Pattern 1: "texte FR / texte EN" (espaces des deux côtés)
    # Pattern 2: "texte FR/ Texte EN" (espace après seulement, EN commence par majuscule)
    # Pattern 3: "texte FR /texte EN" (espace avant seulement)
    match = re.match(r'^(.+?)\s*/\s*([A-Z].+)$', text)
    if match:
        fr = match.group(1).strip()
        en = match.group(2).strip()
        return fr, en

    # Pattern alternatif: "/" avec espace avant mais pas après, suivi de minuscule
    match = re.match(r'^(.+?)\s+/\s*(.+)$', text)
    if match:
        fr = match.group(1).strip()
        en = match.group(2).strip()
        return fr, en

    return text, None


def parse_demarne_fishing_method(
    product_name: Optional[str] = None,
    categorie: Optional[str] = None,
    variante: Optional[str] = None,
    label: Optional[str] = None
) -> Optional[str]:
    """
    Extrait la méthode de pêche depuis les champs Demarne.

    Priorité de recherche : Variante > Label > Categorie > ProductName

    Args:
        product_name: Nom complet du produit
        categorie: Catégorie du produit (ex: "SAUMON SUPÉRIEUR NORVÈGE")
        variante: Variante du produit (ex: "Ligne", "Entier")
        label: Label/certification (ex: "MSC", "BIO")

    Returns:
        Méthode de pêche extraite (ex: "LIGNE", "PB", "IKEJIME") ou None
    """
    # Mots-clés de méthode de pêche à rechercher
    # Ordre important : patterns plus longs d'abord pour éviter les faux positifs
    fishing_methods = [
        ("DE LIGNE", "LIGNE"),
        ("LIGNE", "LIGNE"),
        ("IKEJIME", "IKEJIME"),
        ("IKE", "IKEJIME"),
        ("PB", "PB"),  # Petite Bouche / Pêche à la Bolinche
        ("CASIER", "CASIER"),
        ("FILET", None),  # Exclure FILET car c'est une découpe, pas une méthode
        ("CHALUT", "CHALUT"),
        ("PALANGRE", "PALANGRE"),
        ("FILEYEUR", "FILEYEUR"),
    ]

    def extract_from_text(text: str) -> Optional[str]:
        """Cherche une méthode de pêche dans un texte."""
        if not text:
            return None
        text_upper = text.upper().strip()

        for pattern, method in fishing_methods:
            if method is None:
                continue  # Patterns à ignorer
            # Chercher le pattern comme mot entier (pas dans un autre mot)
            if re.search(rf'\b{pattern}\b', text_upper):
                return method
        return None

    # Chercher dans l'ordre de priorité
    # 1. Variante (plus susceptible de contenir "Ligne", etc.)
    result = extract_from_text(variante)
    if result:
        return result

    # 2. Label (peut contenir des certifications de pêche)
    result = extract_from_text(label)
    if result:
        return result

    # 3. Categorie
    result = extract_from_text(categorie)
    if result:
        return result

    # 4. ProductName en dernier recours
    result = extract_from_text(product_name)
    if result:
        return result

    return None


def propagate_variante_translations(df: pd.DataFrame) -> pd.DataFrame:
    """
    Propage les traductions anglaises des variantes.
    Pour chaque Variante FR sans traduction EN, cherche si une autre ligne
    a la même Variante FR avec une traduction EN et la copie.
    """
    # Construire un dictionnaire des traductions connues: Variante FR -> Variante EN
    translations = {}
    for _, row in df.iterrows():
        variante_fr = row.get("Variante")
        variante_en = row.get("Variante_EN")
        if variante_fr and variante_en:
            translations[variante_fr] = variante_en

    # Propager les traductions
    if translations:
        def fill_variante_en(row):
            if row["Variante_EN"] is None and row["Variante"] in translations:
                return translations[row["Variante"]]
            return row["Variante_EN"]

        df["Variante_EN"] = df.apply(fill_variante_en, axis=1)
        logger.info(f"Demarne: {len(translations)} traductions de variantes propagées")

    return df


def get_merged_cells_map(sheet):
    """
    Construit un dictionnaire des cellules fusionnées.
    Pour chaque cellule (row, col) dans une plage fusionnée,
    retourne la valeur de la cellule en haut à gauche de la plage.

    Returns:
        dict: {(row, col): value} pour toutes les cellules dans des plages fusionnées
    """
    merged_map = {}

    for merged_range in sheet.merged_cells.ranges:
        # Valeur de la cellule en haut à gauche
        top_left_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value

        # Propager cette valeur à toutes les cellules de la plage
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_map[(row, col)] = top_left_value

    return merged_map


def extract_data_from_excel(file_input, date_fallback: Optional[str] = None) -> pd.DataFrame:
    """
    Parse le fichier Excel Demarne avec extraction structurée de toutes les colonnes.
    Gère les cellules fusionnées (notamment Calibre et Origine).
    Accepte soit un chemin de fichier (str), soit le contenu du fichier (bytes).
    Retourne un DataFrame avec colonnes: Code, Categorie, Categorie_EN, Variante,
    Variante_EN, Label, Calibre, Origine, Colisage, Tarif, Unite_Facturee, Date, Vendor
    """
    # Extraction de la date (gère bytes/IO en interne)
    price_date = extract_date_from_excel_header(file_input, date_fallback)

    if price_date is None:
        raise ValueError(
            "ERREUR: Date manquante. "
            "Aucune date n'a été trouvée dans les métadonnées du fichier Excel "
            "et aucun paramètre 'date' n'a été fourni dans la requête."
        )

    # Gérer l'entrée bytes pour openpyxl et pandas
    if isinstance(file_input, bytes):
        wb_input = io.BytesIO(file_input)
        pd_input = io.BytesIO(file_input)  # Créer une nouvelle instance
    else:
        wb_input = file_input
        pd_input = file_input

    # Charger avec openpyxl pour détecter les cellules fusionnées
    wb = load_workbook(wb_input, data_only=True)
    sheet = wb.active

    # Construire la map des cellules fusionnées
    merged_map = get_merged_cells_map(sheet)
    logger.info(f"Demarne: {len(merged_map)} cellules dans des plages fusionnées détectées")

    # Lire aussi avec pandas pour faciliter l'itération
    df_raw = pd.read_excel(pd_input, header=None)

    # Variables de contexte pour héritage (Catégorie et Variante seulement)
    current_categorie = None
    current_categorie_en = None
    current_variante = None
    current_variante_en = None

    # Séparateurs de section à ignorer
    separators = ["LA MARÉE", "LA MAREE", "LES COQUILLAGES"]

    entries = []

    for idx, row in df_raw.iterrows():
        excel_row = idx + 1  # Excel est 1-indexed

        # Fonction helper pour obtenir la valeur (fusionnée ou non)
        def get_cell_value(col_idx):
            excel_col = col_idx + 1  # Excel est 1-indexed
            # Vérifier si la cellule est dans une plage fusionnée
            if (excel_row, excel_col) in merged_map:
                return merged_map[(excel_row, excel_col)]
            # Sinon, utiliser la valeur pandas (None si vide)
            val = row[col_idx]
            return val if pd.notna(val) else None

        # Récupérer les valeurs avec gestion des fusions
        col0_val = get_cell_value(0)
        col1_val = get_cell_value(1)
        col2_val = get_cell_value(2)  # Calibre - peut être fusionné
        col3_val = get_cell_value(3)  # Code
        col4_val = get_cell_value(4)  # Origine - peut être fusionné
        col5_val = get_cell_value(5)  # Colisage
        col6_val = get_cell_value(6)  # Tarif
        col7_val = get_cell_value(7)  # Unité

        # Convertir en strings nettoyées
        col0 = str(col0_val).strip() if col0_val is not None else ""
        col1 = str(col1_val).strip() if col1_val is not None else ""
        col2 = str(col2_val).strip() if col2_val is not None else ""
        col3 = col3_val  # Garder tel quel pour vérifier nan
        col4 = str(col4_val).strip() if col4_val is not None else ""
        col5 = str(col5_val).strip() if col5_val is not None else ""
        col6 = col6_val  # Garder numérique
        col7 = str(col7_val).strip() if col7_val is not None else ""

        # Détecter ligne d'en-tête de paragraphe (contient "Code" ou "Calibre" ou "Caliber" en col 3)
        if col3 in ['Code', 'Calibre', 'Caliber']:
            if col0 and col0 not in separators:
                current_categorie, current_categorie_en = split_fr_en(col0)
                current_variante = None
                current_variante_en = None
            continue

        # Ignorer lignes sans code valide
        if col3 is None:
            continue

        # Vérifier si col3 est un code valide (doit être numérique)
        try:
            code_str = str(col3).strip()
            if code_str in ["", "nan"]:
                continue
            float(code_str)  # Les codes Demarne sont numériques
        except ValueError:
            # Ce n'est pas un code valide, ignorer (probablement un séparateur)
            continue

        # Détecter nouvelle variante (col 0 non vide et ce n'est pas un séparateur)
        if col0 and col0 != "nan" and col0 not in separators:
            current_variante, current_variante_en = split_fr_en(col0)

        # Construire le nom complet du produit
        name_parts = []
        if current_categorie:
            name_parts.append(current_categorie)
        if current_variante:
            name_parts.append(current_variante)
        if col1:  # Label
            name_parts.append(col1)
        if col2:  # Calibre
            name_parts.append(col2)

        product_name = " - ".join(filter(None, name_parts))

        # Convertir le tarif de manière sécurisée
        tarif = None
        if col6 is not None:
            if isinstance(col6, (int, float)) and not pd.isna(col6):
                tarif = float(col6)
            elif isinstance(col6, str):
                try:
                    tarif = float(col6)
                except ValueError:
                    pass

        # Extraire la méthode de pêche
        methode_peche = parse_demarne_fishing_method(
            product_name=product_name,
            categorie=current_categorie,
            variante=current_variante,
            label=col1 if col1 else None
        )

        # Extraire les données
        entry = {
            "Code": code_str,
            "Categorie": current_categorie,
            "Categorie_EN": current_categorie_en,
            "Variante": current_variante,
            "Variante_EN": current_variante_en,
            "Methode_Peche": methode_peche,
            "Label": col1 if col1 else None,
            "Calibre": col2 if col2 else None,
            "Origine": col4 if col4 else None,
            "Colisage": col5 if col5 else None,
            "Tarif": tarif,
            "Unite_Facturee": col7 if col7 else None,
            "ProductName": product_name,
            "Date": price_date,
            "Vendor": "Demarne"
        }
        entries.append(entry)

    # Construire DataFrame
    df = pd.DataFrame(entries)

    # Propager les traductions anglaises des variantes
    df = propagate_variante_translations(df)

    # Ajouter keyDate (Code + Date, sans préfixe vendor)
    df["keyDate"] = df["Code"].astype(str) + "_" + df["Date"].astype(str)
    df["Code_Provider"] = df["Code"].astype(str)  # Alias pour compatibilité, ensure string
    df["Prix"] = df["Tarif"]  # Alias pour compatibilité

    # Supprimer la colonne Code originale (INT64) pour éviter les problèmes de type BigQuery
    df = df.drop(columns=["Code"], errors="ignore")

    logger.info(f"Demarne structured parsing: {len(df)} lignes extraites")

    return df


def parse(file_input, harmonize: bool = False, date_fallback: str = None, **kwargs) -> list[dict]:
    """
    Point d'entrée principal du parser Demarne.

    Extrait les produits d'un fichier Excel Demarne et optionnellement
    applique les règles d'harmonisation pour normaliser les attributs.

    Args:
        file_input: Chemin vers le fichier Excel OU contenu binaire
        harmonize: Si True, applique l'harmonisation des attributs
                   (normalise les valeurs selon docs/harmonisation_attributs.md)
        date_fallback: Date de fallback si non trouvée dans le fichier (format YYYY-MM-DD)
        **kwargs: Arguments supplémentaires (réservés pour extensions futures)

    Returns:
        Liste de dictionnaires produits avec les champs:
        - Date, Vendor, keyDate, Code_Provider, Prix, ProductName
        - Categorie, Variante, Methode_Peche, Label, Calibre, Origine
        - Categorie_EN, Variante_EN (versions anglaises)

        Si harmonize=True, les clés sont en snake_case et les valeurs normalisées:
        - categorie (espèce extraite: SAUMON, BAR, DORADE...)
        - type_production (SAUVAGE, ELEVAGE extrait de la catégorie)
        - qualite (SUP, PREMIUM extrait de la catégorie)
        - etat (ENTIER, CUIT, VIDE extrait de catégorie/variante)
        - decoupe (FILET, DOS, PAVE extrait de variante)
        - origine (normalisée, poids filtrés)
        - label (MSC, BIO, ASC, LABEL ROUGE, IGP, AOP)
        - trim (TRIM_B, TRIM_D, TRIM_E)

    Example:
        >>> products = parse("cours_demarne.xlsx", harmonize=True, date_fallback="2026-01-15")
        >>> products[0]["categorie"]
        'SAUMON'  # Extrait depuis 'SAUMON SUPÉRIEUR NORVÈGE'
    """
    # Extraction des données brutes
    df = extract_data_from_excel(file_input, date_fallback=date_fallback)
    products = sanitize_for_json(df)

    # Application optionnelle de l'harmonisation
    if harmonize:
        if not HARMONIZE_AVAILABLE:
            raise ImportError(
                "Le module services.harmonize n'est pas disponible. "
                "Vérifiez que services/harmonize.py existe."
            )
        products = harmonize_products(products, vendor="Demarne")

    return products


# Alias pour compatibilité
parse_demarne = parse
